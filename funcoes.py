from pathlib import Path


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÃO: detectar_relat
# Objetivo: Abrir o arquivo .xls (salvo como texto separado por tab, UTF-16)
#           e identificar de qual relatório SAP se trata lendo as primeiras
#           linhas do cabeçalho.
#
#   • Relatório A (FBL1N)  → contém a palavra 'Nome' nas 7 primeiras linhas
#                            → cabeçalho real começa na linha 9 (skiprows=8)
#   • Relatório B (REL_LOG) → não contém 'Nome'
#                            → cabeçalho real começa na linha 4 (skiprows=3)
#
# Retorna:
#   df         – DataFrame bruto já lido do arquivo
#   tipo_relat – 'A' para FBL1N  |  'B' para REL_LOG
# ──────────────────────────────────────────────────────────────────────────────
def detectar_relat(caminho_arquivo):

    # Lê apenas as 7 primeiras linhas para inspecionar o cabeçalho SAP
    with open(caminho_arquivo, 'r', encoding='utf-16') as arquivo:
        linhas_iniciais = [arquivo.readline() for _ in range(7)]
        amostra_texto = ''.join(linhas_iniciais)

    # A presença de 'Nome' no cabeçalho indica o layout do FBL1N
    if 'Nome' in amostra_texto:
        pular_linhas = 8       # FBL1N tem 8 linhas de cabeçalho antes dos dados
        tipo_relat = 'A'
    else:
        pular_linhas = 3       # REL_LOG tem 3 linhas de cabeçalho
        tipo_relat = 'B'

    import pandas as pd
    
    # Lê o arquivo completo já pulando as linhas de cabeçalho SAP
    # index_col=False evita que a primeira coluna vire índice (importante no REL_LOG)
    df = pd.read_csv(
        caminho_arquivo,
        encoding='utf-16',
        skiprows=pular_linhas,
        sep='\t',
        index_col=False
    )

    return df, tipo_relat


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÃO: parse_sap_valor
# Objetivo: Tratar de forma robusta os valores monetários do SAP.
# O SAP costuma exportar créditos com o sinal de menos no final (ex: "1.500,00-")
# o que quebra o conversor nativo do Pandas.
# ──────────────────────────────────────────────────────────────────────────────
def parse_sap_valor(val):
    import pandas as pd
    
    if pd.isna(val):
        return 0.0
    val_str = str(val).strip()
    if not val_str:
        return 0.0
        
    is_negative = False
    if val_str.endswith('-'):
        is_negative = True
        val_str = val_str[:-1]
    elif val_str.startswith('-'):
        is_negative = True
        val_str = val_str[1:]
        
    val_str = val_str.replace('.', '').replace(',', '.')
    
    try:
        num = float(val_str)
        if is_negative:
            num = num * -1
        
        return num
    except ValueError:
        return 0.0


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÃO: tratamento_fbl1n
# Objetivo: Limpar e padronizar o DataFrame do relatório FBL1N (Tipo A).
#
# Passos:
#   1. Remove colunas 'Unnamed' (geradas pelo Excel/SAP que exporta colunas vazias)
#   2. Remove colunas desnecessárias pelos índices posicionais [0, 12, 14, 15]
#   3. Renomeia as colunas restantes para nomes amigáveis
#   4. Converte 'Valor' de forma robusta lidando com sinal negativo e multiplicando por -1
#   5. Converte colunas de texto, inteiro e data para os tipos corretos
#   6. Filtra apenas os tipos de documento relevantes: RE, KT, RF
#
# Recebe: df – DataFrame bruto retornado por detectar_relat()
# Retorna: df_filtrado – DataFrame limpo e filtrado
# ──────────────────────────────────────────────────────────────────────────────
def tratamento_fbl1n(df):
    import pandas as pd

    # Tipos de documento que interessam para a conciliação de pagamentos
    puxar_apenas = ['RE', 'KT', 'RF']

    colunas_texto = ['Empresa', 'Referência', 'Texto', 'Tipo']
    colunas_int   = ['LNeg', 'Fornecedor', 'Nº documento', 'Doc. Compensação']
    colunas_data  = ['Data doc.', 'Vencimento', 'Compensaç.']

    # Passo 1 – Remove qualquer coluna cujo nome comece com 'Unnamed'
    df_limpo = df.loc[:, ~df.columns.str.contains('^Unnamed')]

    # Passo 2 – Remove colunas posicionais que não são usadas neste relatório
    df_limpo = df_limpo.drop(df_limpo.columns[[0, 12, 14, 15]], axis=1)

    # Passo 3 – Renomeia as colunas restantes para um padrão único entre relatórios
    colunas_desejadas = [
        'Empresa', 'Fornecedor', 'LNeg', 'Referência', 'Nº documento',
        'Data doc.', 'Doc. Compensação', 'Valor', 'Vencimento', 'Texto',
        'Tipo', 'Compensaç.'
    ]
    df_limpo.columns = colunas_desejadas

    # Passo 4 – Trata o campo 'Valor' com a função robusta
    df_limpo['Valor'] = df_limpo['Valor'].apply(parse_sap_valor) * -1

    # Passo 5 – Converte tipos de cada grupo de colunas
    df_limpo[colunas_texto] = df_limpo[colunas_texto].astype(str)
    df_limpo[colunas_int]   = df_limpo[colunas_int].astype('Int64')

    for coluna in colunas_data:
        df_limpo[coluna] = df_limpo[coluna].astype(str).str.replace('.', '/', regex=False)
        df_limpo[coluna] = pd.to_datetime(df_limpo[coluna], format='%d/%m/%Y', errors='coerce')

    # Passo 6 – Filtra apenas os tipos de documento relevantes
    df_filtrado = df_limpo[df_limpo['Tipo'].isin(puxar_apenas)]

    return df_filtrado


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÃO: tratamento_log
# Objetivo: Limpar e padronizar o DataFrame do relatório REL_LOG (Tipo B).
#
# A lógica é análoga ao tratamento_fbl1n, mas o layout do REL_LOG é diferente:
#   • Colunas a remover: índices [15, 14, 13, 8] (remoção de trás para frente
#     evita que o shift de índices afete as posições seguintes)
#   • O mapa de renomeação é feito por posição (dicionário índice → nome)
#     pois as colunas do REL_LOG não têm nomes limpos como o FBL1N
#
# Recebe: df – DataFrame bruto retornado por detectar_relat()
# Retorna: df_filtrado – DataFrame limpo e filtrado
# ──────────────────────────────────────────────────────────────────────────────
def tratamento_log(df):
    import pandas as pd

    puxar_apenas = ['RE', 'KT', 'RF']

    colunas_texto = ['Empresa', 'Referência', 'Texto', 'Tipo']
    colunas_int   = ['LNeg', 'Fornecedor', 'Nº documento', 'Doc. Compensação']
    colunas_data  = ['Data doc.', 'Vencimento', 'Compensaç.']

    # Passo 1 – Remove colunas 'Unnamed'
    df_limpo = df.loc[:, ~df.columns.str.contains('^Unnamed')]

    # Passo 2 – Remove colunas desnecessárias do layout REL_LOG
    # Os índices são removidos do maior para o menor para não deslocar as posições
    df_limpo = df_limpo.drop(df_limpo.columns[[15, 14, 13, 8]], axis=1)

    # Passo 3 – Renomeia por posição (REL_LOG exporta nomes de coluna diferentes)
    colunas_desejadas = {
        0:  'Empresa',
        1:  'LNeg',
        2:  'Fornecedor',
        3:  'Nº documento',
        4:  'Data doc.',
        5:  'Referência',
        6:  'Vencimento',
        7:  'Tipo',
        8:  'Valor',
        9:  'Texto',
        10: 'Doc. Compensação',
        11: 'Compensaç.'
    }
    dicionario_traduzido = {
        df_limpo.columns[pos]: novo_nome
        for pos, novo_nome in colunas_desejadas.items()
    }
    df_limpo = df_limpo.rename(columns=dicionario_traduzido)

    # Passo 4 – Trata o campo 'Valor' com a função robusta
    df_limpo['Valor'] = df_limpo['Valor'].apply(parse_sap_valor)

    # Passo 5 – Converte tipos
    df_limpo[colunas_texto] = df_limpo[colunas_texto].astype(str)
    df_limpo[colunas_int]   = df_limpo[colunas_int].astype('Int64')

    for coluna in colunas_data:
        df_limpo[coluna] = df_limpo[coluna].astype(str).str.replace('.', '/', regex=False)
        df_limpo[coluna] = pd.to_datetime(df_limpo[coluna], format='%d/%m/%Y', errors='coerce')

    # Passo 6 – Filtra tipos de documento relevantes
    df_filtrado = df_limpo[df_limpo['Tipo'].isin(puxar_apenas)]

    return df_filtrado


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÃO: processar_arquivo  ← ponto de entrada chamado pelo ope.py
# Objetivo: Orquestrar a detecção + tratamento de um único arquivo.
#
#   1. Chama detectar_relat() para identificar o tipo e obter o df bruto
#   2. Direciona para a função de tratamento correta (A → fbl1n | B → log)
#
# Recebe: caminho_arquivo – objeto Path ou str com o caminho do .xls
# Retorna: DataFrame tratado e filtrado
# ──────────────────────────────────────────────────────────────────────────────
def processar_arquivo(caminho_arquivo):

    # Detecta o tipo de relatório e carrega o df bruto em uma única chamada
    df, tipo_relat = detectar_relat(caminho_arquivo)

    if tipo_relat == 'A':
        # Arquivo FBL1N → tratamento específico do layout de extrato de fornecedor
        return tratamento_fbl1n(df)
    else:
        # Arquivo REL_LOG → tratamento do layout de log de pagamentos
        return tratamento_log(df)


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÃO: exportar_excel  [gerado]
# Objetivo: Salvar o df_final em um arquivo .xlsx, criando abas separadas
#           de acordo com a escolha feita pelo usuário na interface, e aplicar
#           formatação visual completa: tabela colorida com zebrado, subtotal
#           de Valor e destaque vermelho para linhas do tipo KT.
#
# Mapeamento de opções da interface → coluna do DataFrame:
#   'Marca'       → coluna 'Empresa'     (nome exibido na UI é diferente do df)
#   'LNeg'        → coluna 'LNeg'
#   'Fornecedor'  → coluna 'Fornecedor'
#   'Não Separar' → sem separação, tudo numa aba 'Dados'
#
# Paletas disponíveis (cor escolhida na interface):
#   Cabeçalho: cor forte  |  Linha ímpar: levemente mais clara  |  Linha par: ainda mais clara
#
# Parâmetros:
#   df_final  – DataFrame consolidado vindo do ope.py
#   separar   – string com a opção escolhida na interface
#   destino   – Path ou str com a pasta de destino
#   nome      – nome base do arquivo (sem extensão)
#   cor       – cor de personalização escolhida na interface: 'Verde' | 'Azul' | 'Roxo' | 'Vermelho'
#
# Retorna: caminho_arquivo (Path) do .xlsx gerado
# ──────────────────────────────────────────────────────────────────────────────
def exportar_excel(df_final, separar, destino, nome, cor='Azul'):
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime
    import tempfile
    import shutil
    import os

    # [gerado] Paletas baseadas nos estilos nativos do Excel "Formatar como Tabela"
    # (TableStyleMedium) com conformidade WCAG AA (contraste ≥ 4.5:1).
    # Estrutura: (cabeçalho, linha ímpar, linha par)
    #   • Cabeçalho: cor sólida escura + texto branco (#FFFFFF) → contraste alto
    #   • Linhas: alternância clara/branca + texto preto (#000000) → máximo legível
    #
    #  Referência Excel:
    #   Azul    → TableStyleMedium2  (#4472C4 / #D9E1F2 / #FFFFFF)
    #   Verde   → TableStyleMedium7  (#70AD47 / #E2EFDA / #FFFFFF)
    #   Roxo    → TableStyleMedium13 (#7030A0 / #E4DFEC / #FFFFFF)
    #   Vermelho→ TableStyleMedium10 (#C00000 / #FFC7CE / #FFFFFF)
    paletas = {
        'Azul':     ('4472C4', 'D9E1F2', 'FFFFFF'),
        'Verde':    ('375623', 'E2EFDA', 'FFFFFF'),
        'Roxo':     ('7030A0', 'E4DFEC', 'FFFFFF'),
        'Vermelho': ('C00000', 'FFC7CE', 'FFFFFF'),
    }

    # [gerado] Seleciona a paleta escolhida; cai em Azul se a cor não for reconhecida
    header_hex, row1_hex, row2_hex = paletas.get(cor, paletas['Azul'])

    fill_header = PatternFill('solid', fgColor=header_hex)
    fill_row1   = PatternFill('solid', fgColor=row1_hex)
    fill_row2   = PatternFill('solid', fgColor=row2_hex)
    fill_kt     = PatternFill('solid', fgColor='FF0000')  # [gerado] KT → vermelho puro

    font_header  = Font(bold=True, color='FFFFFF', size=11)
    font_normal  = Font(color='000000', size=10)
    font_kt      = Font(color='FFFFFF', bold=True, size=10)  # [gerado] texto branco sobre vermelho
    font_subtotal = Font(bold=True, color='FFFFFF', size=10)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left   = Alignment(horizontal='left',   vertical='center')
    align_right  = Alignment(horizontal='right',  vertical='center')

    # [gerado] Bordas brancas para simplicidade visual
    lado_branco = Side(style='thin', color='FFFFFF')
    borda_fina = Border(
        left=lado_branco, right=lado_branco,
        top=lado_branco, bottom=lado_branco,
    )

    # [gerado] Índice posicional da coluna 'Valor' no DataFrame (base 1 para o Excel)
    col_valor_idx = list(df_final.columns).index('Valor') + 1
    col_tipo_idx  = list(df_final.columns).index('Tipo')

    # [gerado] Mapeia o label da UI para o nome real da coluna no DataFrame
    mapa_colunas = {
        'Marca':      'Empresa',
        'LNeg':       'LNeg',
        'Fornecedor': 'Fornecedor',
    }

    # [gerado] Garante que o destino seja um objeto Path
    caminho_arquivo = Path(destino) / f'{nome}.xlsx'

    # [gerado] Cria um arquivo temporário oculto
    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)

    # [gerado] Monta a lista de (nome_aba, DataFrame) para iterar de forma uniforme
    # independente de estar separando por coluna ou não
    if separar == 'Não Separar':
        abas = [('Dados', df_final)]
    else:
        coluna = mapa_colunas[separar]
        abas = [
            (str(valor)[:31], df_final[df_final[coluna] == valor])
            for valor in sorted(df_final[coluna].dropna().unique())
        ]

    # [gerado] Passamos str(caminho_arquivo) em vez do objeto Path.
    # Quando o pandas recebe um Path, em algumas versões ele cria um file-handle 
    # e o openpyxl pode não finalizar/fechar o arquivo ZIP (XLSX) corretamente
    # ao final, gerando o erro de "extensão inválida" ou arquivo corrompido.
    # [gerado] Adicionamos formatação nativa de datetime para garantir dd/mm/yyyy no excel
    with pd.ExcelWriter(
        str(temp_path),
        engine='openpyxl',
        datetime_format='dd/mm/yyyy',
        date_format='dd/mm/yyyy'
    ) as writer:

        for nome_aba, df_aba in abas:

            # [gerado] Grava os dados sem índice; a formatação será aplicada depois
            df_aba.to_excel(writer, sheet_name=nome_aba, index=False)
            ws = writer.sheets[nome_aba]
            n_cols = len(df_aba.columns)

            # ── Formatar cabeçalho (linha 1) ─────────────────────────────────
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill      = fill_header
                cell.font      = font_header
                cell.alignment = align_center
                cell.border    = borda_fina

            # ── Formatar linhas de dados (a partir da linha 2) ────────────────
            for row_idx, row_data in enumerate(df_aba.itertuples(index=False), start=2):

                # [gerado] Verifica se o tipo da linha é KT para pintar de vermelho
                tipo_valor = str(getattr(row_data, 'Tipo', ''))
                is_kt = tipo_valor.strip().upper() == 'KT'

                # [gerado] Zebrado: linhas ímpares e pares recebem tons diferentes
                fill_linha = fill_row1 if row_idx % 2 != 0 else fill_row2

                for col_idx in range(1, n_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = borda_fina

                    if is_kt and col_idx == col_valor_idx:
                        # [gerado] Célula de Valor em linha KT: fundo vermelho + fonte branca
                        cell.fill = fill_kt
                        cell.font = font_kt
                    else:
                        cell.fill = fill_linha
                        cell.font = font_normal

                    # [gerado] Alinhamento por tipo de dado
                    if col_idx == col_valor_idx:
                        cell.alignment = align_right
                    else:
                        cell.alignment = align_left
                        
                    # [gerado] Força a formatação visual brasileira nas datas (dd/mm/yyyy)
                    if isinstance(cell.value, datetime.datetime):
                        cell.number_format = 'dd/mm/yyyy'

            # ── Linha de subtotal de Valor ────────────────────────────────────
            subtotal_row = ws.max_row + 1
            lbl_col = col_valor_idx - 1 if col_valor_idx > 1 else 1  # coluna ao lado esquerdo

            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=subtotal_row, column=col_idx)
                cell.fill   = fill_header
                cell.border = borda_fina
                
                if col_idx == lbl_col:
                    cell.value     = 'SUBTOTAL'
                    cell.font      = font_subtotal
                    cell.alignment = align_right
                elif col_idx == col_valor_idx:
                    cell.value = f'=SUM({get_column_letter(col_valor_idx)}2:' \
                                 f'{get_column_letter(col_valor_idx)}{subtotal_row - 1})'
                    cell.font      = font_subtotal
                    cell.alignment = align_right
                    cell.number_format = '#,##0.00'
                else:
                    cell.font      = font_subtotal

            # ── Ajuste automático da largura das colunas ──────────────────────
            for col_idx in range(1, n_cols + 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ''))
                    for r in range(1, subtotal_row + 1)
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # [gerado] Move atomicamente o temporário completo para o destino final
    # Isso faz o arquivo "nascer" na pasta destino no exato milissegundo em que 
    # fica completamente pronto, sincronizando 100% com o popup da interface.
    shutil.move(temp_path, str(caminho_arquivo))

    return caminho_arquivo